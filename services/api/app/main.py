
from fastapi import FastAPI, UploadFile, BackgroundTasks, WebSocket, WebSocketDisconnect, HTTPException
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse
from fastapi.middleware.cors import CORSMiddleware
from sqlalchemy.ext.asyncio import create_async_engine, async_sessionmaker
from sqlalchemy import select, update, text
from sqlalchemy.orm import DeclarativeBase, Mapped, mapped_column
from sqlalchemy.dialects.postgresql import UUID, JSONB
from datetime import datetime
from typing import List, Optional
import httpx, os, uuid, asyncio, aiofiles, json, logging
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import io

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

app = FastAPI(title="BubbleGrade API", version="1.0.0")

# Add CORS middleware
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Database setup
DATABASE_URL = os.getenv("DATABASE_URL", "postgresql+asyncpg://omr:omr@db/omr")
engine = create_async_engine(DATABASE_URL, echo=True)
async_session = async_sessionmaker(engine, expire_on_commit=False)

class Base(DeclarativeBase):
    pass

class Scan(Base):
    __tablename__ = "scans"
    
    id: Mapped[UUID] = mapped_column(UUID, primary_key=True, default=uuid.uuid4)
    filename: Mapped[str] = mapped_column()
    status: Mapped[str] = mapped_column(default="QUEUED")
    score: Mapped[Optional[int]] = mapped_column()
    answers: Mapped[Optional[dict]] = mapped_column(JSONB)
    total_questions: Mapped[Optional[int]] = mapped_column()
    upload_time: Mapped[datetime] = mapped_column(default=datetime.utcnow)
    processed_time: Mapped[Optional[datetime]] = mapped_column()

# WebSocket connection manager
class ConnectionManager:
    def __init__(self):
        self.active_connections: List[WebSocket] = []

    async def connect(self, websocket: WebSocket):
        await websocket.accept()
        self.active_connections.append(websocket)
        logger.info(f"WebSocket connected. Total connections: {len(self.active_connections)}")

    def disconnect(self, websocket: WebSocket):
        if websocket in self.active_connections:
            self.active_connections.remove(websocket)
        logger.info(f"WebSocket disconnected. Total connections: {len(self.active_connections)}")

    async def broadcast(self, message: dict):
        for connection in self.active_connections.copy():
            try:
                await connection.send_json(message)
            except:
                self.disconnect(connection)

manager = ConnectionManager()

OMR_URL = os.getenv("OMR_URL", "http://omr:8090/grade")

async def call_omr_and_update(scan_path: str, scan_id: str, filename: str):
    try:
        # Update status to PROCESSING
        async with async_session() as session:
            await session.execute(
                update(Scan).where(Scan.id == scan_id).values(status="PROCESSING")
            )
            await session.commit()
        
        await manager.broadcast({
            "type": "scan_update",
            "scan_id": scan_id,
            "status": "PROCESSING"
        })
        
        # Call OMR service
        async with httpx.AsyncClient() as client:
            with open(scan_path, "rb") as f:
                files = {"file": (filename, f, "image/jpeg")}
                resp = await client.post(OMR_URL, files=files, timeout=60.0)
                resp.raise_for_status()
                result = resp.json()
        
        # Update database with results
        async with async_session() as session:
            await session.execute(
                update(Scan).where(Scan.id == scan_id).values(
                    status="COMPLETED",
                    score=result.get("score"),
                    answers=result.get("answers", []),
                    total_questions=result.get("total", 0),
                    processed_time=datetime.utcnow()
                )
            )
            await session.commit()
        
        # Broadcast completion
        await manager.broadcast({
            "type": "scan_complete",
            "scan_id": scan_id,
            "status": "COMPLETED",
            "score": result.get("score"),
            "answers": result.get("answers", [])
        })
        
        logger.info(f"Scan {scan_id} completed with score {result.get('score')}")
        
    except Exception as e:
        logger.error(f"Error processing scan {scan_id}: {e}")
        
        # Update status to ERROR
        async with async_session() as session:
            await session.execute(
                update(Scan).where(Scan.id == scan_id).values(status="ERROR")
            )
            await session.commit()
        
        await manager.broadcast({
            "type": "scan_error",
            "scan_id": scan_id,
            "status": "ERROR",
            "error": str(e)
        })
    
    finally:
        # Clean up temp file
        try:
            os.remove(scan_path)
        except:
            pass

@app.websocket("/ws")
async def websocket_endpoint(websocket: WebSocket):
    await manager.connect(websocket)
    try:
        while True:
            await websocket.receive_text()
    except WebSocketDisconnect:
        manager.disconnect(websocket)

@app.post("/api/scans")
async def upload_scan(file: UploadFile, tasks: BackgroundTasks):
    scan_id = str(uuid.uuid4())
    dest = f"/tmp/{scan_id}_{file.filename}"
    
    # Save file
    async with aiofiles.open(dest, "wb") as out:
        content = await file.read()
        await out.write(content)
    
    # Create database record
    async with async_session() as session:
        scan = Scan(
            id=scan_id,
            filename=file.filename,
            status="QUEUED"
        )
        session.add(scan)
        await session.commit()
    
    # Start background processing
    tasks.add_task(call_omr_and_update, dest, scan_id, file.filename)
    
    return {"id": scan_id, "status": "QUEUED", "filename": file.filename}

@app.get("/api/scans")
async def list_scans():
    async with async_session() as session:
        result = await session.execute(select(Scan).order_by(Scan.upload_time.desc()))
        scans = result.scalars().all()
        return [
            {
                "id": str(scan.id),
                "filename": scan.filename,
                "status": scan.status,
                "score": scan.score,
                "answers": scan.answers,
                "total_questions": scan.total_questions,
                "upload_time": scan.upload_time.isoformat(),
                "processed_time": scan.processed_time.isoformat() if scan.processed_time else None
            }
            for scan in scans
        ]

@app.get("/api/scans/{scan_id}")
async def get_scan(scan_id: str):
    async with async_session() as session:
        result = await session.execute(select(Scan).where(Scan.id == scan_id))
        scan = result.scalar_one_or_none()
        
        if not scan:
            raise HTTPException(status_code=404, detail="Scan not found")
        
        return {
            "id": str(scan.id),
            "filename": scan.filename,
            "status": scan.status,
            "score": scan.score,
            "answers": scan.answers,
            "total_questions": scan.total_questions,
            "upload_time": scan.upload_time.isoformat(),
            "processed_time": scan.processed_time.isoformat() if scan.processed_time else None
        }

@app.get("/api/exports/{scan_id}")
async def export_scan(scan_id: str):
    async with async_session() as session:
        result = await session.execute(select(Scan).where(Scan.id == scan_id))
        scan = result.scalar_one_or_none()
        
        if not scan:
            raise HTTPException(status_code=404, detail="Scan not found")
        
        if scan.status != "COMPLETED":
            raise HTTPException(status_code=400, detail="Scan not completed yet")
    
    # Create Excel file
    wb = Workbook()
    ws = wb.active
    ws.title = "Bubble Sheet Results"
    
    # Headers
    headers = ["Question", "Answer", "Correct Answer", "Result"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # Correct answers (for demo - would be configurable in real app)
    correct_answers = ["A", "B", "C", "D", "A", "B", "C", "D", "A", "B"]
    answers = scan.answers or []
    
    # Data rows
    for i, (answer, correct) in enumerate(zip(answers, correct_answers), 1):
        ws.cell(row=i+1, column=1, value=f"Q{i}")
        ws.cell(row=i+1, column=2, value=answer)
        ws.cell(row=i+1, column=3, value=correct)
        
        result_cell = ws.cell(row=i+1, column=4)
        if answer == correct:
            result_cell.value = "✓"
            result_cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
        else:
            result_cell.value = "✗"
            result_cell.fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")
    
    # Summary
    summary_row = len(answers) + 3
    ws.cell(row=summary_row, column=1, value="Summary").font = Font(bold=True)
    ws.cell(row=summary_row+1, column=1, value="Score:")
    ws.cell(row=summary_row+1, column=2, value=f"{scan.score}%")
    ws.cell(row=summary_row+2, column=1, value="Correct:")
    ws.cell(row=summary_row+2, column=2, value=f"{sum(1 for a, c in zip(answers, correct_answers) if a == c)}/{len(correct_answers)}")
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to memory
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return FileResponse(
        path="/dev/stdin",
        filename=f"{scan.filename}_{scan_id}_results.xlsx",
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        content=output.getvalue()
    )

@app.get("/health")
async def health_check():
    try:
        async with async_session() as session:
            await session.execute(text("SELECT 1"))
        return {"status": "healthy", "service": "api", "database": "connected"}
    except Exception as e:
        return {"status": "unhealthy", "service": "api", "database": "disconnected", "error": str(e)}

@app.on_event("startup")
async def startup():
    logger.info("🚀 BubbleGrade API starting up...")
    logger.info(f"📊 OMR service URL: {OMR_URL}")
    logger.info(f"🗄️ Database URL: {DATABASE_URL}")

@app.on_event("shutdown")
async def shutdown():
    logger.info("🛑 BubbleGrade API shutting down...")
    await engine.dispose()
