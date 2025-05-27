import httpx
import os
import aiofiles
import uuid
from typing import List
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from ..domain.entities import OMRResult, WebSocketMessage
from ..domain.services import OMRService, FileStorageService, ExcelExportService


class HttpOMRService(OMRService):
    def __init__(self):
        self.omr_url = os.getenv("OMR_URL", "http://omr:8090/grade")

    async def process_image(self, file_path: str, filename: str) -> OMRResult:
        async with httpx.AsyncClient() as client:
            with open(file_path, "rb") as f:
                files = {"file": (filename, f, "image/jpeg")}
                response = await client.post(self.omr_url, files=files, timeout=60.0)
                response.raise_for_status()
                result = response.json()
                
                return OMRResult(
                    score=result.get("score", 0),
                    answers=result.get("answers", []),
                    total_questions=result.get("total", 0)
                )


class LocalFileStorageService(FileStorageService):
    def __init__(self, base_path: str = "/tmp"):
        self.base_path = base_path

    async def save_upload(self, file_content: bytes, filename: str) -> str:
        file_id = str(uuid.uuid4())
        file_path = f"{self.base_path}/{file_id}_{filename}"
        
        async with aiofiles.open(file_path, "wb") as f:
            await f.write(file_content)
        
        return file_path

    async def delete_file(self, file_path: str) -> None:
        try:
            os.remove(file_path)
        except OSError:
            pass  # File might already be deleted


class OpenpyxlExcelExportService(ExcelExportService):
    async def create_report(self, scan_id: str, filename: str, score: int, 
                           answers: List[str], correct_answers: List[str]) -> bytes:
        wb = Workbook()
        ws = wb.active
        ws.title = "Bubble Sheet Results"
        
        # Headers
        headers = ["Question", "Answer", "Correct Answer", "Result"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Data rows
        for i, (answer, correct) in enumerate(zip(answers, correct_answers), 1):
            ws.cell(row=i+1, column=1, value=f"Q{i}")
            ws.cell(row=i+1, column=2, value=answer)
            ws.cell(row=i+1, column=3, value=correct)
            
            result_cell = ws.cell(row=i+1, column=4)
            if answer == correct:
                result_cell.value = "✓"
                result_cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
            else:
                result_cell.value = "✗"
                result_cell.fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")
        
        # Summary
        summary_row = len(answers) + 3
        ws.cell(row=summary_row, column=1, value="Summary").font = Font(bold=True)
        ws.cell(row=summary_row+1, column=1, value="Score:")
        ws.cell(row=summary_row+1, column=2, value=f"{score}%")
        ws.cell(row=summary_row+2, column=1, value="Correct:")
        correct_count = sum(1 for a, c in zip(answers, correct_answers) if a == c)
        ws.cell(row=summary_row+2, column=2, value=f"{correct_count}/{len(correct_answers)}")
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to memory
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()